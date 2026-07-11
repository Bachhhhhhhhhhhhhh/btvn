# -*- coding: utf-8 -*-
"""Extract LOG depreciation Excel → rich data.json for Doraemon dashboard v3."""
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime, date
import json

DOWNLOADS = Path(r"C:\Users\User\Downloads")
CANDIDATES = [
    DOWNLOADS / "Bản Log chính thức_PREMIUM.xlsx",
    DOWNLOADS / "Bản Log chính thức.xlsx",
    DOWNLOADS / "Bản Log chính thức_CONG_THUC.xlsx",
]
xlsx_path = next((p for p in CANDIDATES if p.exists()), None)
if not xlsx_path:
    raise SystemExit("Excel source not found")

wb = load_workbook(xlsx_path, data_only=True)

GL_NAMES = {
    6424001: "Structure",
    6424002: "Building",
    6424004: "Equipment",
    6424005: "Vehicle",
    6424009: "Other / CIP",
}
CC_NAMES = {
    91501: "LOG – SGA",
    91541: "LOG – FAC3 (Hà Nam)",
    13801: "CC 13801",
    15101: "MCP → LOG",
}


def n(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return 0.0


def dstr(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m")
    if isinstance(v, date):
        return v.strftime("%Y-%m")
    return None


def safe_int(v):
    try:
        return None if v is None else int(float(v))
    except Exception:
        return None


def parse_ym(s):
    if not s:
        return None
    try:
        return datetime.strptime(s + "-01", "%Y-%m-%d")
    except Exception:
        return None


# Budget MT
ws = wb["Budget MT27-31"]
hdr_r = 2
for r in range(1, 6):
    if ws.cell(r, 2).value == "CC" or ws.cell(r, 3).value == "G/L":
        hdr_r = r
        break
periods = ["103Ki", "104Ki", "105Ki", "106Ki", "107Ki", "108Ki"]
col_map = {4: "103Ki", 5: "104Ki", 7: "105Ki", 9: "106Ki", 11: "107Ki", 13: "108Ki"}
gap_map = {6: "104Ki", 8: "105Ki", 10: "106Ki", 12: "107Ki", 14: "108Ki"}

budget_rows = []
totals_by_period = {p: 0.0 for p in periods}
for r in range(hdr_r + 1, ws.max_row + 1):
    cc, gl = safe_int(ws.cell(r, 2).value), safe_int(ws.cell(r, 3).value)
    if cc is None or gl is None:
        continue
    row = {
        "cc": cc,
        "ccName": CC_NAMES.get(cc, str(cc)),
        "gl": gl,
        "glName": GL_NAMES.get(gl, str(gl)),
        "values": {},
        "gaps": {},
    }
    for c, pname in col_map.items():
        val = n(ws.cell(r, c).value)
        row["values"][pname] = val
        totals_by_period[pname] += val
    for c, pname in gap_map.items():
        row["gaps"][pname] = n(ws.cell(r, c).value)
    budget_rows.append(row)

grand = {}
for c, pname in col_map.items():
    v = n(ws.cell(1, c).value)
    grand[pname] = v if v else totals_by_period[pname]

# Monthly 103
ws = wb["Budget 103Ki 1QFC"]
months = [dstr(ws.cell(2, c).value) or f"M{c}" for c in range(4, 16)]
monthly_rows = []
monthly_total = [0.0] * len(months)
for r in range(3, ws.max_row + 1):
    cc, gl = safe_int(ws.cell(r, 2).value), safe_int(ws.cell(r, 3).value)
    if cc is None:
        continue
    vals = [n(ws.cell(r, c).value) for c in range(4, 4 + len(months))]
    for i, v in enumerate(vals):
        monthly_total[i] += v
    monthly_rows.append(
        {
            "cc": cc,
            "ccName": CC_NAMES.get(cc, str(cc)),
            "gl": gl,
            "glName": GL_NAMES.get(gl, str(gl)),
            "months": vals,
            "total": n(ws.cell(r, 16).value) or sum(vals),
        }
    )

# Assets
ws = wb["MT27-31 Depreciation file"]
assets = []
today = datetime(2026, 7, 11)
for r in range(7, ws.max_row + 1):
    name = ws.cell(r, 12).value
    code = ws.cell(r, 11).value
    cost = ws.cell(r, 13).value
    cc = ws.cell(r, 7).value or ws.cell(r, 2).value
    gl = ws.cell(r, 20).value
    life = ws.cell(r, 18).value
    start = ws.cell(r, 17).value
    end = ws.cell(r, 19).value
    atype = ws.cell(r, 21).value
    if not name and not code:
        continue
    period_tots = [
        n(ws.cell(r, c).value) if c <= ws.max_column else 0.0
        for c in [36, 50, 64, 78, 92, 106]
    ]
    cc_i, gl_i = safe_int(cc), safe_int(gl)
    end_s, start_s = dstr(end), dstr(start)
    end_dt = parse_ym(end_s)
    months_left = None
    status = "active"
    if end_dt:
        months_left = (end_dt.year - today.year) * 12 + (end_dt.month - today.month)
        if months_left < 0:
            status = "ended"
        elif months_left <= 24:
            status = "ending_soon"
        else:
            status = "active"
    monthly_rate = (n(cost) / n(life)) if n(life) else 0.0
    assets.append(
        {
            "code": str(code) if code else "",
            "name": str(name) if name else "",
            "cost": n(cost),
            "cc": cc_i,
            "ccName": CC_NAMES.get(cc_i, str(cc_i) if cc_i else "—"),
            "gl": gl_i,
            "glName": GL_NAMES.get(gl_i, str(gl_i) if gl_i else "—"),
            "life": n(life) if life else None,
            "monthlyRate": monthly_rate,
            "start": start_s,
            "end": end_s,
            "monthsLeft": months_left,
            "status": status,
            "assetType": str(atype) if atype else "",
            "periodTotals": period_tots,
            "totalAll": sum(period_tots),
        }
    )

assets_sorted = sorted(assets, key=lambda x: x["cost"], reverse=True)

by_cc, by_gl = {}, {}
for row in budget_rows:
    by_cc.setdefault(row["ccName"], {p: 0.0 for p in periods})
    by_gl.setdefault(row["glName"], {p: 0.0 for p in periods})
    for p in periods:
        by_cc[row["ccName"]][p] += row["values"].get(p, 0)
        by_gl[row["glName"]][p] += row["values"].get(p, 0)

# Waterfall
wf, running = [], 0.0
base = grand.get("103Ki") or totals_by_period["103Ki"]
wf.append({"label": "103Ki", "value": base, "type": "total"})
running = base
for i in range(1, len(periods)):
    prev, cur = periods[i - 1], periods[i]
    delta = (grand.get(cur) or totals_by_period[cur]) - (
        grand.get(prev) or totals_by_period[prev]
    )
    wf.append({"label": f"Δ {cur}", "value": delta, "type": "delta"})
    running += delta
    wf.append({"label": cur, "value": running, "type": "total"})

heatmap = [
    {
        "key": f"{r['ccName']} · {r['glName']}",
        "cc": r["ccName"],
        "gl": r["glName"],
        "values": [r["values"].get(p, 0) for p in periods],
    }
    for r in budget_rows
]

growth = {}
for i in range(1, len(periods)):
    prev = grand.get(periods[i - 1]) or totals_by_period[periods[i - 1]] or 1
    cur = grand.get(periods[i]) or totals_by_period[periods[i]]
    growth[periods[i]] = (cur - prev) / prev if prev else 0

# Top movers by absolute gap 104 vs 103
movers = []
for r in budget_rows:
    gap = r["gaps"].get("104Ki", (r["values"].get("104Ki", 0) - r["values"].get("103Ki", 0)))
    movers.append(
        {
            "key": f"{r['ccName']} · {r['glName']}",
            "ccName": r["ccName"],
            "glName": r["glName"],
            "gap": gap,
            "from": r["values"].get("103Ki", 0),
            "to": r["values"].get("104Ki", 0),
        }
    )
movers_up = sorted([m for m in movers if m["gap"] > 0], key=lambda x: -x["gap"])[:8]
movers_down = sorted([m for m in movers if m["gap"] < 0], key=lambda x: x["gap"])[:8]

status_counts = {"active": 0, "ending_soon": 0, "ended": 0}
for a in assets:
    status_counts[a["status"]] = status_counts.get(a["status"], 0) + 1

# Insights text for helper
insights = []
peak = max(periods, key=lambda p: grand.get(p) or totals_by_period[p])
insights.append(f"Kỳ cao nhất là {peak} với {grand.get(peak) or totals_by_period[peak]:,.0f} VND.")
if movers_up:
    insights.append(
        f"Tăng mạnh nhất 104vs103: {movers_up[0]['key']} (+{movers_up[0]['gap']:,.0f})."
    )
if movers_down:
    insights.append(
        f"Giảm mạnh nhất 104vs103: {movers_down[0]['key']} ({movers_down[0]['gap']:,.0f})."
    )
if status_counts["ending_soon"]:
    insights.append(
        f"Có {status_counts['ending_soon']} tài sản sắp hết khấu hao trong 24 tháng."
    )
insights.append(f"Đang theo dõi {len(assets)} dòng tài sản · nguyên giá ~{sum(a['cost'] for a in assets)/1e9:.1f}B VND.")

data = {
    "meta": {
        "title": "Doraemon LOG Depreciation",
        "unit": "VND",
        "source": xlsx_path.name,
        "generated": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "periods": periods,
        "months103": months,
        "version": "3.0-dora",
    },
    "kpi": {
        "total103": grand.get("103Ki") or totals_by_period["103Ki"],
        "totalMT": sum(totals_by_period.values()),
        "assetCount": len(assets),
        "totalAcquisition": sum(a["cost"] for a in assets),
        "avgMonthly103": (grand.get("103Ki") or totals_by_period["103Ki"]) / 12,
        "peakPeriod": peak,
        "byPeriod": totals_by_period,
        "grand": grand if any(grand.values()) else totals_by_period,
        "growth": growth,
        "statusCounts": status_counts,
    },
    "budgetRows": budget_rows,
    "byCC": by_cc,
    "byGL": by_gl,
    "waterfall": wf,
    "heatmap": heatmap,
    "moversUp": movers_up,
    "moversDown": movers_down,
    "insights": insights,
    "monthly": {"labels": months, "total": monthly_total, "rows": monthly_rows},
    "assets": assets_sorted[:50],
    "assetsAll": assets_sorted,
    "glNames": {str(k): v for k, v in GL_NAMES.items()},
    "ccNames": {str(k): v for k, v in CC_NAMES.items()},
}

out = Path(__file__).resolve().parent / "data.json"
out.write_text(json.dumps(data, ensure_ascii=False), encoding="utf-8")
print("OK", out.name, out.stat().st_size, "assets", len(assets), "insights", len(insights))
